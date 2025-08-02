#!/usr/bin/env python3
"""
Field analysis script for multimarketplace upload system.

This script analyzes all input files to identify:
- Dropdown fields and their values
- Field name variations across platforms
- Data validation rules
- Field types (dropdown vs free text)
"""

import sys
import pandas as pd
import openpyxl
import json
from pathlib import Path
from typing import Dict, List, Set, Tuple, Any
from collections import defaultdict
import logging

# Add src to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from src.config.settings import INPUT_FILES, ensure_directories_exist

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

class FieldAnalyzer:
    """Analyzes fields across all marketplace files."""
    
    def __init__(self):
        self.analysis_results = {
            'files_analyzed': [],
            'dropdown_fields': {},
            'field_variations': defaultdict(list),
            'data_validation_rules': {},
            'field_types': {},
            'unique_values_per_field': {},
            'platform_specific_fields': {}
        }
    
    def analyze_excel_file(self, file_path: Path) -> Dict[str, Any]:
        """Analyze a single Excel file for fields and dropdowns."""
        file_analysis = {
            'file_name': file_path.name,
            'headers': [],
            'dropdown_fields': {},
            'field_variations': {},
            'data_validation': {},
            'unique_values': {},
            'field_types': {}
        }
        
        try:
            logger.info(f"Analyzing {file_path.name}...")
            
            # Read headers
            df = pd.read_excel(file_path, nrows=1)
            file_analysis['headers'] = df.columns.tolist()
            
            # Read sample data for analysis
            df_sample = pd.read_excel(file_path, nrows=50)  # First 50 rows
            
            # Analyze each column
            for column in df.columns:
                self._analyze_column(file_path, column, df_sample, file_analysis)
            
            # Check for data validation (dropdowns)
            self._check_data_validation(file_path, file_analysis)
            
            logger.info(f"Completed analysis of {file_path.name}")
            return file_analysis
            
        except Exception as e:
            logger.error(f"Error analyzing {file_path}: {e}")
            return file_analysis
    
    def _analyze_column(self, file_path: Path, column: str, df_sample: pd.DataFrame, file_analysis: Dict):
        """Analyze a single column for field characteristics."""
        try:
            # Get unique values (excluding NaN)
            unique_values = df_sample[column].dropna().unique()
            unique_count = len(unique_values)
            
            # Determine field type
            field_type = self._determine_field_type(column, unique_values, unique_count)
            file_analysis['field_types'][column] = field_type
            
            # Store unique values (limit to first 20 for analysis)
            file_analysis['unique_values'][column] = {
                'count': unique_count,
                'sample_values': unique_values[:20].tolist(),
                'type': field_type
            }
            
            # Check for dropdown characteristics
            if field_type == 'dropdown_candidate':
                file_analysis['dropdown_fields'][column] = {
                    'type': 'dropdown_candidate',
                    'unique_count': unique_count,
                    'sample_values': unique_values[:10].tolist()
                }
            
            # Identify field variations
            normalized_name = self._normalize_field_name(column)
            file_analysis['field_variations'][normalized_name] = {
                'original_name': column,
                'normalized_name': normalized_name,
                'field_type': field_type
            }
            
        except Exception as e:
            logger.error(f"Error analyzing column {column} in {file_path.name}: {e}")
    
    def _determine_field_type(self, column: str, unique_values: List, unique_count: int) -> str:
        """Determine the type of field based on characteristics."""
        column_lower = column.lower()
        
        # Check for known dropdown fields
        dropdown_indicators = [
            'colour', 'color', 'couleur', 'kolor', 'kleur',
            'material', 'matériau', 'materiał', 'materiaal',
            'category', 'catégorie', 'kategoria', 'categorie',
            'brand', 'marque', 'marka', 'merk',
            'weight unit', 'unit', 'unité', 'jednostka',
            'dimension unit', 'measurement unit'
        ]
        
        for indicator in dropdown_indicators:
            if indicator in column_lower:
                if unique_count <= 50:  # Likely dropdown if limited unique values
                    return 'dropdown_candidate'
                else:
                    return 'dropdown_free'  # Dropdown with custom entries
        
        # Check for free text fields
        text_indicators = [
            'title', 'titre', 'tytuł',
            'description', 'descripción', 'descrizione',
            'name', 'nom', 'nazwa', 'naam',
            'usp', 'feature', 'caractéristique'
        ]
        
        for indicator in text_indicators:
            if indicator in column_lower:
                return 'free_text'
        
        # Default classification based on unique value count
        if unique_count <= 20:
            return 'dropdown_candidate'
        elif unique_count <= 100:
            return 'dropdown_free'
        else:
            return 'free_text'
    
    def _normalize_field_name(self, field_name: str) -> str:
        """Normalize field names for comparison across platforms."""
        # Common field name mappings
        field_mappings = {
            # Colour variations
            'couleur': 'colour',
            'kolor': 'colour', 
            'kleur': 'colour',
            'color': 'colour',
            
            # Material variations
            'matériau': 'material',
            'materiał': 'material',
            'materiaal': 'material',
            
            # Category variations
            'catégorie': 'category',
            'kategoria': 'category',
            'categorie': 'category',
            
            # Brand variations
            'marque': 'brand',
            'marka': 'brand',
            'merk': 'brand',
            
            # Weight variations
            'poids': 'weight',
            'waga': 'weight',
            'gewicht': 'weight',
            
            # Dimension variations
            'dimension': 'dimension',
            'wymiar': 'dimension',
            'afmeting': 'dimension'
        }
        
        normalized = field_name.lower().strip()
        
        # Apply mappings
        for original, standard in field_mappings.items():
            if original in normalized:
                return standard
        
        return normalized
    
    def _check_data_validation(self, file_path: Path, file_analysis: Dict):
        """Check for Excel data validation (dropdowns)."""
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            worksheet = workbook.active
            
            validation_fields = {}
            
            # Check first 20 rows for data validation
            for row in range(2, min(22, worksheet.max_row + 1)):
                for col in range(1, min(50, worksheet.max_column + 1)):  # Check first 50 columns
                    cell = worksheet.cell(row=row, column=col)
                    
                    if cell.data_validation:
                        # Get the column header
                        header_cell = worksheet.cell(row=1, column=col)
                        header = header_cell.value
                        
                        if header and header not in validation_fields:
                            validation_fields[header] = {
                                'validation_type': str(cell.data_validation.type),
                                'formula1': str(cell.data_validation.formula1) if cell.data_validation.formula1 else None,
                                'formula2': str(cell.data_validation.formula2) if cell.data_validation.formula2 else None,
                                'allow_blank': cell.data_validation.allowBlank,
                                'show_error_message': cell.data_validation.showErrorMessage,
                                'show_input_message': cell.data_validation.showInputMessage
                            }
            
            file_analysis['data_validation'] = validation_fields
            
        except Exception as e:
            logger.error(f"Error checking data validation in {file_path}: {e}")
    
    def analyze_all_files(self) -> Dict[str, Any]:
        """Analyze all input files."""
        logger.info("Starting comprehensive field analysis...")
        
        input_dir = Path("data/input")
        
        for file_path in input_dir.glob("*.xlsx"):
            if file_path.name.startswith("~$"):  # Skip temporary files
                continue
                
            file_analysis = self.analyze_excel_file(file_path)
            self.analysis_results['files_analyzed'].append(file_analysis)
            
            # Aggregate results
            self._aggregate_results(file_analysis)
        
        # Generate summary statistics
        self._generate_summary()
        
        return self.analysis_results
    
    def _aggregate_results(self, file_analysis: Dict):
        """Aggregate results from individual file analysis."""
        file_name = file_analysis['file_name']
        
        # Aggregate dropdown fields
        for field, details in file_analysis['dropdown_fields'].items():
            if field not in self.analysis_results['dropdown_fields']:
                self.analysis_results['dropdown_fields'][field] = []
            self.analysis_results['dropdown_fields'][field].append({
                'file': file_name,
                'details': details
            })
        
        # Aggregate field variations
        for normalized_name, details in file_analysis['field_variations'].items():
            self.analysis_results['field_variations'][normalized_name].append({
                'file': file_name,
                'original_name': details['original_name'],
                'field_type': details['field_type']
            })
        
        # Aggregate data validation
        self.analysis_results['data_validation_rules'][file_name] = file_analysis['data_validation']
        
        # Aggregate unique values
        self.analysis_results['unique_values_per_field'][file_name] = file_analysis['unique_values']
    
    def _generate_summary(self):
        """Generate summary statistics."""
        total_files = len(self.analysis_results['files_analyzed'])
        total_dropdown_fields = len(self.analysis_results['dropdown_fields'])
        total_field_variations = len(self.analysis_results['field_variations'])
        
        logger.info(f"Analysis complete: {total_files} files, {total_dropdown_fields} dropdown fields, {total_field_variations} field variations")
    
    def save_results(self, output_path: Path = None):
        """Save analysis results to JSON file."""
        if output_path is None:
            output_path = Path("data/analysis_results.json")
        
        ensure_directories_exist()
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(self.analysis_results, f, indent=2, ensure_ascii=False, default=str)
        
        logger.info(f"Analysis results saved to {output_path}")
    
    def print_summary(self):
        """Print a human-readable summary of the analysis."""
        print("\n" + "=" * 80)
        print("📊 FIELD ANALYSIS SUMMARY")
        print("=" * 80)
        
        # Files analyzed
        print(f"\n📁 Files Analyzed: {len(self.analysis_results['files_analyzed'])}")
        for file_analysis in self.analysis_results['files_analyzed']:
            print(f"   • {file_analysis['file_name']}: {len(file_analysis['headers'])} fields")
        
        # Dropdown fields
        print(f"\n🔽 Dropdown Fields Found: {len(self.analysis_results['dropdown_fields'])}")
        for field, occurrences in self.analysis_results['dropdown_fields'].items():
            files = [occ['file'] for occ in occurrences]
            print(f"   • {field}: {', '.join(files)}")
        
        # Field variations
        print(f"\n🔄 Field Name Variations: {len(self.analysis_results['field_variations'])}")
        for normalized_name, variations in self.analysis_results['field_variations'].items():
            if len(variations) > 1:  # Only show fields that vary
                print(f"   • {normalized_name}:")
                for var in variations:
                    print(f"     - {var['file']}: '{var['original_name']}' ({var['field_type']})")
        
        # Data validation
        print(f"\n✅ Data Validation Rules: {len(self.analysis_results['data_validation_rules'])} files")
        for file_name, rules in self.analysis_results['data_validation_rules'].items():
            if rules:
                print(f"   • {file_name}: {len(rules)} validation rules")
        
        print("\n" + "=" * 80)

def main():
    """Main function to run field analysis."""
    print("🔍 Starting comprehensive field analysis...")
    print("=" * 60)
    
    # Ensure directories exist
    ensure_directories_exist()
    
    # Create analyzer and run analysis
    analyzer = FieldAnalyzer()
    results = analyzer.analyze_all_files()
    
    # Save results
    analyzer.save_results()
    
    # Print summary
    analyzer.print_summary()
    
    print("\n✅ Field analysis completed successfully!")
    print("📝 Next steps:")
    print("1. Review analysis_results.json for detailed findings")
    print("2. Create field mapping configurations")
    print("3. Implement dropdown standardization")
    print("4. Update template optimizer with dropdown support")

if __name__ == "__main__":
    main() 