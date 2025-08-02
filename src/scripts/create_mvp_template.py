#!/usr/bin/env python3
"""
MVP Template Creator

Creates a minimal template with only the required fields across all platforms.
This allows for fast product uploads with minimal data entry.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from pathlib import Path
import logging

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# MVP Required Fields (merged duplicates)
MVP_FIELDS = [
    # Core identifiers (all platforms)
    'Category Code',        # Maps to Category, Product Category (auto-populates Product Category)
    'EAN',                 # All platforms require this (auto-populates Shop SKU, Product Identifier)
    'Product Title (Mirakl)',  # Maps to Name, Product Title FR/IT/ES/PT
    'Description (Mirakl)',    # Maps to Body Copy, Long Description (auto-populates Body Copy)
    'Image 1',            # Maps to Main Image 1, Image Large 1 (auto-populates Main Image 1 and Image Large 1)
    'Brand',              # Maps to Brand Name, Acquisition brand, Manufacturer Name (auto-populates all brand fields)
    
    # Castorama specific fields
    'Pack quantity',      # Castorama_PL
    'Pack type',          # Castorama_PL
    'Product type',       # Castorama_PL
    'Product thickness (mm)', # Castorama_PL, Castorama_FR
    'Product width (mm)', # Castorama_FR
    'Colour group',       # Castorama_FR
    'Effect',             # Castorama_FR
    'Location',           # Castorama_FR
    'Facet_product_type', # Castorama_FR
    
    # Maxeda specific fields
    'Seller Product ID',  # Maxeda_BE, Maxeda_NL
    'Product Title (nl_BE)', # Maxeda_BE
    'Product Title (fr_BE)', # Maxeda_BE
    'Description (nl_BE)', # Maxeda_BE
    'Description (fr_BE)', # Maxeda_BE
    'Image 2',            # Maxeda_BE, Maxeda_NL
    'USP 1 nl_be',       # Maxeda_BE
    'USP 2 nl_be',       # Maxeda_BE
    'USP 3 nl_be',       # Maxeda_BE
    'USP 1 fr_be',       # Maxeda_BE
    'USP 2 fr_be',       # Maxeda_BE
    'USP 3 fr_be',       # Maxeda_BE
    'Material',           # Maxeda_BE, Maxeda_NL
    'Product Title (nl_NL)', # Maxeda_NL
    'Description (nl_NL)', # Maxeda_NL
    'USP 1 (nl_NL)',     # Maxeda_NL
    'USP 2 (nl_NL)',     # Maxeda_NL
    'USP 3 (nl_NL)',     # Maxeda_NL
    
    # LM specific fields
    'Product Title FR',    # LM
    'Product Title IT',    # LM
    'Product Title ES',    # LM
    'Product Title PT',    # LM
    'Long Description FR', # LM
    'Long Description IT', # LM
    'Long Description ES', # LM
    'Long Description PT', # LM
    'Contains wood'        # LM
]

# Auto-population mappings (simple local mappings)
AUTO_POPULATE_MAPPINGS = {
    # EAN auto-populates Shop SKU and Product Identifier
    'Shop SKU': 'EAN',
    'Product Identifier': 'EAN',
    
    # Category Code auto-populates Product Category
    'Product Category': 'Category Code',
    
    # Description (Mirakl) auto-populates Body Copy
    'Body Copy': 'Description (Mirakl)',
    
    # Image 1 auto-populates Main Image 1 and Image Large 1
    'Main Image 1': 'Image 1',
    'Image Large 1': 'Image 1',
    
    # Brand auto-populates all brand fields
    'Acquisition brand': 'Brand',
    'Brand Name': 'Brand',
    'Manufacturer Name': 'Brand',
    
    # Guarantee fields are merged
    'Guarantee': 'Manufacturer guarantee',
    'Guarantee exclusion specifications': 'Manufacturer guarantee'
}

def create_mvp_template():
    """Create an MVP template with only the minimum required fields."""
    
    print("🎯 Creating MVP Template (Minimum Required Fields)")
    print("=" * 60)
    
    # Create empty DataFrame with MVP fields
    df = pd.DataFrame(columns=MVP_FIELDS)
    
    # Create output path
    output_path = Path('data/templates/multimarketplace_mvp_template.xlsx')
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    try:
        # Save as XLSX with styling
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='MVP Template', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['MVP Template']
            
            # Apply styling
            _apply_mvp_styling(worksheet, MVP_FIELDS)
            _adjust_mvp_column_widths(worksheet)
            
        logger.info(f"✅ Created MVP template with {len(MVP_FIELDS)} fields")
        logger.info(f"📁 Output: {output_path}")
        
        print(f"\n✅ MVP Template Created Successfully!")
        print(f"📁 File: {output_path}")
        print(f"📊 Fields: {len(MVP_FIELDS)} required fields (merged duplicates)")
        
        print("\n🎨 Color Coding:")
        print("   🔴 Red: Required fields (must be filled)")
        print("   🟢 Green: Translation fields (auto-translated if blank)")
        
        print("\n🔄 Auto-Population Features:")
        print("   • EAN → Shop SKU, Product Identifier (automatic)")
        print("   • Category Code → Product Category (automatic)")
        print("   • Description (Mirakl) → Body Copy (automatic)")
        print("   • Image 1 → Main Image 1, Image Large 1 (automatic)")
        print("   • Brand → Acquisition brand, Brand Name, Manufacturer Name (automatic)")
        print("   • Guarantee fields merged (automatic)")
        
        print("\n🌐 Translation Support:")
        print("   • Green fields can be left blank for auto-translation")
        print("   • Fill manually if you have translations to save API usage")
        print("   • DeepL will translate from Product Title (Mirakl) and Description (Mirakl)")
        
        print("\n📋 Platform Coverage:")
        print("   • Castorama FR: 9 unique fields (8 auto-populated)")
        print("   • Castorama PL: 5 unique fields (8 auto-populated)")
        print("   • Maxeda BE: 12 unique fields (7 auto-populated)")
        print("   • Maxeda NL: 8 unique fields (7 auto-populated)")
        print("   • LM: 8 unique fields (7 auto-populated)")
        
        print("\n📝 Next Steps:")
        print("1. Fill in required fields (red) for each product")
        print("2. Leave translation fields (green) blank for auto-translation")
        print("3. Auto-populated fields will be filled automatically")
        print("4. Run transformation: python main.py transform")
        print("5. Upload generated files to respective marketplaces")
        
        print("\n💡 Simple Local Tool:")
        print("   • No client configuration needed")
        print("   • Direct field mappings")
        print("   • Fast client onboarding")
        
        return True
        
    except Exception as e:
        logger.error(f"❌ Error creating MVP template: {e}")
        print(f"❌ Error: {e}")
        return False

def _apply_mvp_styling(worksheet, fields):
    """Apply color coding to MVP template."""
    # Colors
    required_color = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')  # Light red
    translation_color = PatternFill(start_color='E6FFE6', end_color='E6FFE6', fill_type='solid')  # Light green
    
    # Translation fields (can be auto-translated)
    translation_fields = [
        'Product Title (nl_BE)', 'Product Title (fr_BE)', 'Product Title (nl_NL)',
        'Description (nl_BE)', 'Description (fr_BE)', 'Description (nl_NL)',
        'USP 1 nl_be', 'USP 2 nl_be', 'USP 3 nl_be',
        'USP 1 fr_be', 'USP 2 fr_be', 'USP 3 fr_be',
        'USP 1 (nl_NL)', 'USP 2 (nl_NL)', 'USP 3 (nl_NL)',
        'Product Title FR', 'Product Title IT', 'Product Title ES', 'Product Title PT',
        'Long Description FR', 'Long Description IT', 'Long Description ES', 'Long Description PT'
    ]
    
    # Apply color coding based on field type
    for col_idx, field in enumerate(fields, 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        
        if field in translation_fields:
            cell.fill = translation_color  # Green - translatable
        else:
            cell.fill = required_color     # Red - required fields

def _adjust_mvp_column_widths(worksheet):
    """Adjust column widths for MVP template."""
    # Define column widths based on field type
    column_widths = {
        # Core fields
        'A': 15,  # Category Code
        'B': 15,  # EAN
        'C': 40,  # Product Title (Mirakl)
        'D': 60,  # Description (Mirakl)
        'E': 20,  # Image 1
        'F': 15,  # Brand
        
        # Castorama fields
        'G': 15,  # Pack quantity
        'H': 15,  # Pack type
        'I': 20,  # Product type
        'J': 20,  # Product thickness (mm)
        'K': 20,  # Product width (mm)
        'L': 15,  # Colour group
        'M': 15,  # Effect
        'N': 15,  # Location
        'O': 25,  # Facet_product_type
        
        # Maxeda fields
        'P': 20,  # Seller Product ID
        'Q': 40,  # Product Title (nl_BE)
        'R': 40,  # Product Title (fr_BE)
        'S': 60,  # Description (nl_BE)
        'T': 60,  # Description (fr_BE)
        'U': 20,  # Image 2
        'V': 30,  # USP 1 nl_be
        'W': 30,  # USP 2 nl_be
        'X': 30,  # USP 3 nl_be
        'Y': 30,  # USP 1 fr_be
        'Z': 30,  # USP 2 fr_be
        'AA': 30, # USP 3 fr_be
        'AB': 20, # Material
        'AC': 40, # Product Title (nl_NL)
        'AD': 60, # Description (nl_NL)
        'AE': 30, # USP 1 (nl_NL)
        'AF': 30, # USP 2 (nl_NL)
        'AG': 30, # USP 3 (nl_NL)
        
        # LM fields
        'AH': 40, # Product Title FR
        'AI': 40, # Product Title IT
        'AJ': 40, # Product Title ES
        'AK': 40, # Product Title PT
        'AL': 60, # Long Description FR
        'AM': 60, # Long Description IT
        'AN': 60, # Long Description ES
        'AO': 60, # Long Description PT
        'AP': 15  # Contains wood
    }
    
    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width

def main():
    """Main function to create MVP template."""
    success = create_mvp_template()
    
    if success:
        print("\n🎉 MVP Template ready for use!")
    else:
        print("\n❌ Failed to create MVP template")

if __name__ == "__main__":
    main() 