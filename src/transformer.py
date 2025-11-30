"""
Core transformation logic - minimalistic and simple.
Maps input data to platform-specific Excel files based on mapping configuration.
"""

import pandas as pd
import shutil
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from pathlib import Path
from typing import Dict, List, Optional
import logging

from .config import (
    MAPPING_FILE, ATTRIBUTES_FILE, REF_TEMPLATES,
    INPUT_DIR, OUTPUT_DIR, get_output_file_path
)

logger = logging.getLogger(__name__)


def load_mapping() -> Dict[str, Dict[str, str]]:
    """
    Load column mappings from mapping.xlsx.
    
    Returns:
        Dict structure: {attribute_name: {platform: column_letter}}
        Example: {'height': {'castorama_fr': 'BM', 'leroy_merlin': 'AU'}}
    """
    try:
        df = pd.read_excel(MAPPING_FILE, sheet_name='Column_Mappings')
        
        # Get all platform columns (everything ending with '_column')
        platform_columns = [col for col in df.columns if col.endswith('_column')]
        
        mapping = {}
        for _, row in df.iterrows():
            attribute = str(row['attribute_name']).strip()
            value_type = str(row['input_type']).strip()  # Still called input_type in Excel, but represents value_type
            data_source = str(row['data_source']).strip() if 'data_source' in row and pd.notna(row.get('data_source')) else 'input_file'
            
            platform_mapping = {}
            for col in platform_columns:
                # Extract platform name from column (e.g., 'castorama_fr_column' -> 'castorama_fr')
                platform = col.replace('_column', '')
                column_letter = str(row[col]).strip() if pd.notna(row[col]) else ''
                
                if column_letter:
                    platform_mapping[platform] = column_letter
            
            mapping[attribute] = {
                'value_type': value_type,  # 'same' or 'platform_specific'
                'data_source': data_source,  # 'input_file' or 'punch_card'
                'columns': platform_mapping
            }
        
        logger.info(f"Loaded {len(mapping)} attribute mappings")
        return mapping
        
    except Exception as e:
        logger.error(f"Error loading mapping file: {e}")
        return {}


def load_attributes() -> Dict[str, Dict]:
    """
    Load pre-determined attribute values from attributes.xlsx.
    
    Returns:
        Dict with two keys:
        - 'same': {attribute_name: value} for same-input attributes
        - 'platform_specific': {attribute_name: {platform: value}} for platform-specific attributes
    """
    try:
        attributes = {
            'same': {},
            'platform_specific': {}
        }
        
        # Load same-input attributes
        same_df = pd.read_excel(ATTRIBUTES_FILE, sheet_name='Same_Input_Attributes')
        for _, row in same_df.iterrows():
            attr_name = str(row['attribute_name']).strip()
            value = row['value']
            if pd.notna(value):
                attributes['same'][attr_name] = value
        
        # Load platform-specific attributes
        platform_df = pd.read_excel(ATTRIBUTES_FILE, sheet_name='Platform_Specific_Attributes')
        for _, row in platform_df.iterrows():
            attr_name = str(row['attribute_name']).strip()
            
            # Get all platform value columns
            platform_values = {}
            for col in platform_df.columns:
                if col.endswith('_value') and col != 'attribute_name':
                    platform = col.replace('_value', '')
                    value = row[col]
                    if pd.notna(value) and str(value).strip():
                        platform_values[platform] = str(value).strip()
            
            if platform_values:
                attributes['platform_specific'][attr_name] = platform_values
        
        logger.info(f"Loaded {len(attributes['same'])} same-input attributes")
        logger.info(f"Loaded {len(attributes['platform_specific'])} platform-specific attributes")
        return attributes
        
    except Exception as e:
        logger.error(f"Error loading attributes file: {e}")
        return {'same': {}, 'platform_specific': {}}


def load_input(input_file: str) -> pd.DataFrame:
    """
    Load product input file.
    
    Args:
        input_file: Filename in data/input/ directory (e.g., 'acoustic_panels.xlsx')
    
    Returns:
        DataFrame with one row per product
    """
    try:
        input_path = INPUT_DIR / input_file
        df = pd.read_excel(input_path)
        logger.info(f"Loaded {len(df)} products from {input_file}")
        return df
    except Exception as e:
        logger.error(f"Error loading input file {input_file}: {e}")
        return pd.DataFrame()


def copy_reference_template(platform: str, output_path: Path) -> bool:
    """
    Copy reference template exactly to output location.
    
    Args:
        platform: Platform identifier (e.g., 'castorama_fr')
        output_path: Path where output file should be saved
    
    Returns:
        True if successful, False otherwise
    """
    try:
        ref_path = REF_TEMPLATES.get(platform)
        if not ref_path or not ref_path.exists():
            logger.error(f"Reference template not found for {platform}: {ref_path}")
            return False
        
        # Copy reference template exactly to output location
        shutil.copy2(ref_path, output_path)
        
        logger.info(f"Copied reference template for {platform} to {output_path}")
        return True
        
    except Exception as e:
        logger.error(f"Error copying reference template for {platform}: {e}")
        return False


def transform_product(
    product_row: pd.Series,
    platform: str,
    mapping: Dict,
    attributes: Dict
) -> Dict[str, any]:
    """
    Transform one product row to platform-specific values.
    
    Args:
        product_row: Single row from input DataFrame
        platform: Target platform identifier
        mapping: Column mapping dictionary from load_mapping()
        attributes: Attribute values from load_attributes()
    
    Returns:
        Dict mapping column names/indices to values: {column_name: value}
    """
    result = {}
    
    for attribute, attr_config in mapping.items():
        value_type = attr_config['value_type']  # 'same' or 'platform_specific'
        data_source = attr_config['data_source']  # 'input_file' or 'punch_card'
        columns = attr_config['columns']
        
        # Get target column for this platform
        if platform not in columns:
            continue
        
        column_letter = columns[platform]
        
        # Determine value based on data_source and value_type
        value = None
        
        if data_source == 'punch_card':
            # Get value from attributes file (punch card)
            if value_type == 'same':
                if attribute in attributes['same']:
                    value = attributes['same'][attribute]
            elif value_type == 'platform_specific':
                if attribute in attributes['platform_specific']:
                    platform_values = attributes['platform_specific'][attribute]
                    if platform in platform_values:
                        value = platform_values[platform]
        
        elif data_source == 'input_file':
            # Get value from input file (product row)
            if attribute in product_row.index and pd.notna(product_row[attribute]):
                value = product_row[attribute]
            else:
                # Handle typo: mapping might have 'desciption_fr' but input has 'description_fr'
                # Try alternative spelling
                if attribute == 'desciption_fr' and 'description_fr' in product_row.index:
                    if pd.notna(product_row['description_fr']):
                        value = product_row['description_fr']
                elif attribute == 'description_fr' and 'desciption_fr' in product_row.index:
                    if pd.notna(product_row['desciption_fr']):
                        value = product_row['desciption_fr']
        
        # Skip if no value found
        if value is None:
            continue
        
        # Store with column letter as key (will be converted to column name/index later)
        result[column_letter] = value
    
    return result


def column_letter_to_index(column_letter: str) -> int:
    """
    Convert Excel column letter (e.g., 'A', 'BM', 'AA') to 0-based index.
    
    Args:
        column_letter: Excel column letter(s)
    
    Returns:
        0-based column index
    """
    column_letter = column_letter.upper().strip()
    result = 0
    for char in column_letter:
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result - 1  # Convert to 0-based


def generate_platform_file(platform: str, input_file: str) -> Optional[Path]:
    """
    Generate platform-specific output file by copying reference template and inserting values.
    
    Simple approach:
    1. Copy reference template exactly to output (preserves all formatting)
    2. Use openpyxl to write values directly to cells (product 1 = row 3, headers = rows 1-2)
    3. Insert values based on mapping
    
    Args:
        platform: Platform identifier
        input_file: Input filename (e.g., 'acoustic_panels.xlsx')
    
    Returns:
        Path to generated file, or None if failed
    """
    try:
        # Load all configurations
        mapping = load_mapping()
        attributes = load_attributes()
        input_df = load_input(input_file)
        
        if input_df.empty:
            logger.warning(f"No input data to transform")
            return None
        
        # Get output path
        output_path = get_output_file_path(platform)
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        
        # Step 1: Copy reference template exactly to output (preserves formatting)
        if not copy_reference_template(platform, output_path):
            return None
        
        # Step 2: Load with openpyxl to preserve formatting
        wb = load_workbook(output_path)
        ws = wb.active  # Work with first sheet
        
        # Step 3: Find and clear existing data rows (keep headers in rows 1-2)
        # Find last row with data
        last_data_row = 2  # Start after header rows (1 and 2)
        for row in range(3, ws.max_row + 1):
            if any(cell.value is not None for cell in ws[row]):
                last_data_row = row
        
        # Delete existing data rows (keep rows 1-2 as headers)
        if last_data_row > 2:
            ws.delete_rows(3, last_data_row - 2)
        
        # Step 4: Transform each product and write directly to cells
        # Product 1 starts at row 3 (rows 1-2 are headers)
        next_row = 3
        for idx, product_row in input_df.iterrows():
            # Get column mappings for this product (returns {column_letter: value})
            column_values = transform_product(product_row, platform, mapping, attributes)
            
            # Auto-map EAN to shop_sku if EAN exists in input and shop_sku is mapped
            if 'ean' in product_row.index and pd.notna(product_row['ean']):
                if 'shop_sku' in mapping and platform in mapping['shop_sku']['columns']:
                    shop_sku_col = mapping['shop_sku']['columns'][platform]
                    column_values[shop_sku_col] = product_row['ean']
                    logger.debug(f"Auto-mapped EAN to shop_sku column {shop_sku_col}")
            
            # Handle typo: mapping has 'desciption_fr' but input might have 'description_fr'
            # Check if desciption_fr is mapped but description_fr exists in input
            if 'desciption_fr' in mapping and platform in mapping['desciption_fr']['columns']:
                desciption_col = mapping['desciption_fr']['columns'][platform]
                # Try both spellings
                if 'description_fr' in product_row.index and pd.notna(product_row['description_fr']):
                    column_values[desciption_col] = product_row['description_fr']
                elif 'desciption_fr' in product_row.index and pd.notna(product_row['desciption_fr']):
                    column_values[desciption_col] = product_row['desciption_fr']
            
            # Write values directly to cells (preserves formatting from reference template)
            for col_letter, value in column_values.items():
                try:
                    # Convert Excel column letter to 1-based column index (openpyxl uses 1-based)
                    col_idx = column_index_from_string(col_letter)
                    
                    # Write value to cell (this preserves cell formatting from reference template)
                    cell = ws.cell(row=next_row, column=col_idx)
                    cell.value = value
                    
                except Exception as e:
                    logger.warning(f"Error writing to column {col_letter} at row {next_row}: {e}")
            
            next_row += 1
        
        # Step 5: Save workbook (preserves all formatting, sheets, etc.)
        wb.save(output_path)
        
        logger.info(f"Generated {output_path} with {next_row - 3} product rows")
        return output_path
        
    except Exception as e:
        logger.error(f"Error generating platform file for {platform}: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return None

