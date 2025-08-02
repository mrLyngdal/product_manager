"""
Template optimizer for creating user-friendly XLSX master template.

This module:
- Converts CSV template to XLSX format
- Restructures columns in logical order
- Applies color coding for field types
- Adds dropdown validation for standardized fields
- Makes template easier for manual data entry
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path
from typing import List, Dict, Tuple
import logging

from ..config.settings import get_template_file_path, TEMPLATES_DIR
from ..config.field_mappings import get_all_dropdown_fields, get_dropdown_values, is_dropdown_field, is_auto_filled_field
from ..config.client_config import get_client_config

logger = logging.getLogger(__name__)

# Color coding for different field types
COLORS = {
    'required': 'FFE6E6',      # Light red - Required fields
    'automated': 'E6FFE6',     # Light green - Automated fields
    'optional': 'FFFFE6',      # Light yellow - Optional fields
    'header': 'CCCCCC'         # Gray - Headers
}

# Column categories and their colors
COLUMN_CATEGORIES = {
    'required': [
        'Category Code',
        'EAN',
        'Code for internal use',
        'Product Title (Mirakl)',
        'Description (Mirakl)',
        'Category',
        'Product weight (kg)',
        'Product height (mm)',
        'Product width (mm)',
        'Product length (mm)',
        'Material',
        'Colour',
        'Image 1'
    ],
    'automated': [
        # Client-specific brand fields (auto-filled)
        'Brand', 'Brand Name', 'Manufacturer Name',
        # Language-specific titles
        'Product Title (fr_BE)', 'Product Title (nl_BE)', 'Product Title (nl_NL)',
        'Product Title ES', 'Product Title FR', 'Product Title IT', 'Product Title PT',
        # Language-specific descriptions
        'Description (fr_BE)', 'Description (nl_BE)', 'Description (nl_NL)',
        'Short Description (fr_BE)', 'Short Description (nl_BE)', 'Short Description (nl_NL)',
        'Long Description ES', 'Long Description FR', 'Long Description IT', 'Long Description PT',
        # Language-specific USPs
        'USP 1 (fr_BE)', 'USP 1 (nl_BE)', 'USP 1 (nl_NL)',
        'USP 2 (fr_BE)', 'USP 2 (nl_BE)', 'USP 2 (nl_NL)',
        'USP 3 (fr_BE)', 'USP 3 (nl_BE)', 'USP 3 (nl_NL)',
        'USP 4 (fr_BE)', 'USP 4 (nl_BE)', 'USP 4 (nl_NL)',
        'USP 5 (fr_BE)', 'USP 5 (nl_BE)', 'USP 5 (nl_NL)',
        # Additional images
        'Image 2', 'Image 3', 'Image 4', 'Image 5', 'Image 6', 'Image 7', 'Image 8', 'Image 9', 'Image 10',
        'Image Large 1', 'Image Large 2', 'Image Large 3', 'Image Large 4', 'Image Large 5',
        'Image Large 6', 'Image Large 7', 'Image Large 8', 'Image Large 9',
        'Secondary Image 1', 'Secondary Image 2', 'Secondary Image 3', 'Secondary Image 4',
        'Secondary Image 5', 'Secondary Image 6', 'Secondary Image 7', 'Secondary Image 8',
        'Main Image 1', 'Back of Pack Image'
    ]
}

# Column order for restructured template
COLUMN_ORDER = [
    # Core identifiers
    'Category Code',
    'EAN',
    'Brand',
    'Code for internal use',
    'Category',
    
    # English content (base language)
    'Product Title (Mirakl)',
    'Description (Mirakl)',
    'Unique Selling Point 1',
    'Unique Selling Point 2',
    'Unique Selling Point 3',
    'Unique Selling Point 4',
    'Unique Selling Point 5',
    'Unique Selling Point 6',
    'Unique Selling Point 7',
    'Unique Selling Point 8',
    
    # French content
    'Product Title (fr_BE)',
    'Description (fr_BE)',
    'Short Description (fr_BE)',
    'USP 1 (fr_BE)',
    'USP 2 (fr_BE)',
    'USP 3 (fr_BE)',
    'USP 4 (fr_BE)',
    'USP 5 (fr_BE)',
    
    # Dutch content
    'Product Title (nl_BE)',
    'Description (nl_BE)',
    'Short Description (nl_BE)',
    'USP 1 (nl_BE)',
    'USP 2 (nl_BE)',
    'USP 3 (nl_BE)',
    'USP 4 (nl_BE)',
    'USP 5 (nl_BE)',
    
    # Dutch NL content
    'Product Title (nl_NL)',
    'Description (nl_NL)',
    'Short Description (nl_NL)',
    'USP 1 (nl_NL)',
    'USP 2 (nl_NL)',
    'USP 3 (nl_NL)',
    'USP 4 (nl_NL)',
    'USP 5 (nl_NL)',
    
    # Spanish content
    'Product Title ES',
    'Long Description ES',
    
    # French content
    'Product Title FR',
    'Long Description FR',
    
    # Italian content
    'Product Title IT',
    'Long Description IT',
    
    # Portuguese content
    'Product Title PT',
    'Long Description PT',
    
    # Physical specifications
    'Product weight (kg)',
    'Product height (mm)',
    'Product width (mm)',
    'Product length (mm)',
    'Product weight (g)',
    'Product thickness (mm)',
    'Product type',
    'Material',
    'Colour',
    'Main Color',
    'Main Material',
    
    # Images
    'Image 1',
    'Image 2',
    'Image 3',
    'Image 4',
    'Image 5',
    'Image 6',
    'Image 7',
    'Image 8',
    'Image 9',
    'Image 10',
    'Image Large 1',
    'Image Large 2',
    'Image Large 3',
    'Image Large 4',
    'Image Large 5',
    'Image Large 6',
    'Image Large 7',
    'Image Large 8',
    'Image Large 9',
    'Secondary Image 1',
    'Secondary Image 2',
    'Secondary Image 3',
    'Secondary Image 4',
    'Secondary Image 5',
    'Secondary Image 6',
    'Secondary Image 7',
    'Secondary Image 8',
    'Main Image 1',
    'Back of Pack Image',
    
    # Product attributes
    'Can be cut',
    'Cut to size',
    'Contains wood',
    'Compatible with damp rooms or bathrooms',
    'Suitable for damp spaces indicator',
    'Usable in humid spaces',
    'Resistant to limescale',
    'Resistant to moisture',
    'Resistant to mould',
    'Resistant to rust',
    'Resistant to ultraviolet',
    'Fire-retardant indicator',
    'CE marked',
    'UKCA marked',
    'FSC Mark indicator',
    'PEFC mark indicator',
    
    # Additional fields (all remaining columns)
]

class TemplateOptimizer:
    """Optimizes the master template for better usability."""
    
    def __init__(self, client_name: str = 'Nordic Acoustics'):
        self.workbook = None
        self.worksheet = None
        self.client_config = get_client_config(client_name)
    
    def load_csv_template(self) -> pd.DataFrame:
        """Load the existing CSV template."""
        # Try to find the CSV template first
        csv_path = TEMPLATES_DIR / "multimarketplace_master_template.csv"
        if not csv_path.exists():
            raise FileNotFoundError(f"Master template not found: {csv_path}")
        
        df = pd.read_csv(csv_path)
        logger.info(f"Loaded CSV template with {len(df.columns)} columns")
        return df
    
    def restructure_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """Restructure columns in the desired order."""
        # Get all columns from the original template
        all_columns = list(df.columns)
        
        # Create new column order
        new_order = []
        
        # Add columns in the specified order
        for col in COLUMN_ORDER:
            if col in all_columns:
                new_order.append(col)
                all_columns.remove(col)
        
        # Add remaining columns at the end
        new_order.extend(all_columns)
        
        # Reorder the dataframe
        df_reordered = df[new_order]
        
        logger.info(f"Restructured template with {len(df_reordered.columns)} columns")
        return df_reordered
    
    def create_xlsx_template(self, df: pd.DataFrame, output_path: Path) -> bool:
        """Create an XLSX template with color coding and dropdown validation."""
        try:
            # Create workbook and worksheet
            self.workbook = openpyxl.Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = "Product Template"
            
            # Add data to worksheet
            for r in dataframe_to_rows(df, index=False, header=True):
                self.worksheet.append(r)
            
            # Apply color coding
            self._apply_color_coding()
            
            # Add dropdown validation
            self._add_dropdown_validation()
            
            # Auto-adjust column widths
            self._adjust_column_widths()
            
            # Save the workbook
            self.workbook.save(output_path)
            logger.info(f"Created optimized XLSX template: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error creating XLSX template: {e}")
            return False
    
    def _apply_color_coding(self):
        """Apply color coding to cells based on field type."""
        # Get header row
        header_row = 1
        
        for col_idx, cell in enumerate(self.worksheet[header_row], 1):
            column_name = cell.value
            
            # Determine color based on field type
            color = self._get_column_color(column_name)
            
            # Apply fill color to header cell
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            
            # Apply bold font to header
            cell.font = Font(bold=True)
    
    def _get_column_color(self, column_name: str) -> str:
        """Get color for a column based on its category."""
        if column_name in COLUMN_CATEGORIES['required']:
            return COLORS['required']
        elif column_name in COLUMN_CATEGORIES['automated']:
            return COLORS['automated']
        else:
            return COLORS['optional']
    
    def _adjust_column_widths(self):
        """Auto-adjust column widths for better readability."""
        for column in self.worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set column width (with some padding)
            adjusted_width = min(max_length + 2, 50)
            self.worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _add_dropdown_validation(self):
        """Add dropdown validation to fields that should have dropdowns."""
        dropdown_fields = get_all_dropdown_fields()
        
        for col_idx, cell in enumerate(self.worksheet[1], 1):  # Header row
            column_name = cell.value
            
            # Skip auto-filled fields (like Brand, Brand Name, Manufacturer Name)
            if is_auto_filled_field(column_name):
                logger.info(f"Skipping dropdown validation for auto-filled field '{column_name}'")
                continue
            
            if column_name in dropdown_fields:
                dropdown_values = get_dropdown_values(column_name)
                
                if dropdown_values:
                    # Create data validation
                    validation = DataValidation(
                        type="list",
                        formula1=f'"{",".join(dropdown_values)}"',
                        allow_blank=True
                    )
                    
                    # Apply validation to the entire column (excluding header)
                    column_letter = cell.column_letter
                    validation.add(f"{column_letter}2:{column_letter}1000")  # Apply to data rows
                    
                    # Add validation to worksheet
                    self.worksheet.add_data_validation(validation)
                    
                    logger.info(f"Added dropdown validation for field '{column_name}' with {len(dropdown_values)} options")
                else:
                    logger.warning(f"No dropdown values found for field '{column_name}'")
    
    def create_sample_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """Add sample data to the template with client-specific brand names."""
        # Create sample rows
        sample_data = []
        
        # Get client-specific brand names
        brand_mappings = self.client_config.get_all_brand_mappings()
        
        # Sample product 1
        sample_row = {}
        for col in df.columns:
            if 'Category Code' in col:
                sample_row[col] = 'FLOOR-001'
            elif 'EAN' in col:
                sample_row[col] = '1234567890123'
            elif 'Brand Name' in col:
                sample_row[col] = brand_mappings.get('Maxeda_BE', 'Nordic Acoustics')
            elif 'Manufacturer Name' in col:
                sample_row[col] = brand_mappings.get('Castorama_FR', 'Nordic Acoustics')
            elif 'Brand' in col and 'Brand Name' not in col and 'Manufacturer Name' not in col:
                # Auto-fill brand field with client-specific brand
                sample_row[col] = brand_mappings.get('LM_product', 'NORDIC ACOUSTICS')
            elif 'Code for internal use' in col:
                sample_row[col] = 'FLOOR-001'
            elif 'Product Title (Mirakl)' in col:
                sample_row[col] = 'Premium Wood Flooring'
            elif 'Description (Mirakl)' in col:
                sample_row[col] = 'High-quality oak wood flooring with natural finish'
            elif 'Category' in col:
                sample_row[col] = 'Flooring'
            elif 'Product weight (kg)' in col:
                sample_row[col] = 15.5
            elif 'Product height (mm)' in col:
                sample_row[col] = 15
            elif 'Product width (mm)' in col:
                sample_row[col] = 200
            elif 'Product length (mm)' in col:
                sample_row[col] = 2000
            elif 'Material' in col:
                sample_row[col] = 'Solid Oak'
            elif 'Colour' in col:
                sample_row[col] = 'Natural Oak'
            elif 'Image 1' in col:
                sample_row[col] = 'https://example.com/images/1234567890123_1.jpg'
            elif 'Can be cut' in col:
                sample_row[col] = 'Yes'
            elif 'Contains wood' in col:
                sample_row[col] = 'Yes'
            elif 'CE marked' in col:
                sample_row[col] = 'Yes'
            else:
                sample_row[col] = ''
        
        sample_data.append(sample_row)
        
        # Create DataFrame with sample data
        sample_df = pd.DataFrame(sample_data)
        
        logger.info(f"Added sample data with client-specific brand names for {self.client_config.get_client_name()}")
        return sample_df
    
    def optimize_template(self) -> bool:
        """Main method to optimize the template."""
        try:
            # Load existing CSV template
            df = self.load_csv_template()
            
            # Restructure columns
            df_restructured = self.restructure_columns(df)
            
            # Create output path
            output_path = TEMPLATES_DIR / "multimarketplace_master_template.xlsx"
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
            # Create XLSX template
            success = self.create_xlsx_template(df_restructured, output_path)
            
            if success:
                logger.info("Template optimization completed successfully")
                logger.info(f"Output file: {output_path}")
                logger.info("Color coding:")
                logger.info(f"  - Red: Required fields ({len(COLUMN_CATEGORIES['required'])} columns)")
                logger.info(f"  - Green: Automated fields ({len(COLUMN_CATEGORIES['automated'])} columns)")
                logger.info(f"  - Yellow: Optional fields (remaining columns)")
            
            return success
            
        except Exception as e:
            logger.error(f"Error optimizing template: {e}")
            return False

def optimize_master_template(client_name: str = 'Nordic Acoustics'):
    """Main function to optimize the master template."""
    optimizer = TemplateOptimizer(client_name)
    return optimizer.optimize_template()

if __name__ == "__main__":
    import logging
    logging.basicConfig(level=logging.INFO)
    optimize_master_template() 